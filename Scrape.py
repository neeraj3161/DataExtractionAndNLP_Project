import locale
import os
import string
import time
import requests
from bs4 import BeautifulSoup
import openpyxl
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize, sent_tokenize
nltk.download('stopwords')
nltk.download('punkt')

# Common variables
extracted_data_folder = "extracted_data"
positive_words_file_name = "positive-words.txt"
negative_words_file_name = "negative-words.txt"
output_file_name = "Output Data Structure.xlsx"
input_file_name = "Input.xlsx"
vowels = "aeiouyAEIOUY"
personal_pronouns = ["i", "we", "my", "ours", "us"]
output_file_headers = [
            "URL_ID", "URL", "POSITIVE SCORE", "NEGATIVE SCORE", "POLARITY SCORE",
            "SUBJECTIVITY SCORE", "AVG SENTENCE LENGTH", "PERCENTAGE OF COMPLEX WORDS",
            "FOG INDEX", "AVG NUMBER OF WORDS PER SENTENCE", "COMPLEX WORD COUNT",
            "WORD COUNT", "SYLLABLE PER WORD", "PERSONAL PRONOUNS", "AVG WORD LENGTH"
        ]
class InputFileDataModel:
    def __init__(self, urlId, url):
        self.urlId = urlId,
        self.url = url

class OutputFileModel:
    def __init__(self, urlId, url,  positive_score, negative_score, polarity_score, subjectivity_score,
                 average_sentence_length, percentage_complex_words, fog_index,
                 average_words_per_sentence, complex_word_count, word_count,
                 syllable_per_word, personal_pronouns, average_word_length):
        self.urlId = urlId
        self.url = url
        self.positive_score = positive_score
        self.negative_score = negative_score
        self.polarity_score = polarity_score
        self.subjectivity_score = subjectivity_score
        self.average_sentence_length = average_sentence_length
        self.percentage_complex_words = percentage_complex_words
        self.fog_index = fog_index
        self.average_words_per_sentence = average_words_per_sentence
        self.complex_word_count = complex_word_count
        self.word_count = word_count
        self.syllable_per_word = syllable_per_word
        self.personal_pronouns = personal_pronouns
        self.average_word_length = average_word_length


def ExtractData(raw_data):
    soup = BeautifulSoup(''.join(raw_data), features="html.parser")
    article_data = soup.find_all("div", class_ = ['td-post-content', 'tagdiv-type'])
    extracted_text = ''
    for para in article_data:
        text = para.get_text(separator='\n', strip=True)
        extracted_text += text
    return extracted_text

def CreateFolder(folderName):
    try:
        os.mkdir(folderName)
        print(f"Folder {folderName} created successfully!!")
    except FileExistsError as error:
        print(f"Folder {folderName} already exists!!")
    except Exception as e:
        print(f"There was an error creating folder {e}")
        quit()

def CreateTxtFileAndWriteData(fileName, data):
     try:
        file = open(os.path.join(os.getcwd(),f"{extracted_data_folder}/{fileName}.txt"), "w", encoding="utf-8")
        file.write(data)
        file.close()
        print(f"File {fileName}.txt created successfully in applications {extracted_data_folder} folder.")
     except Exception as e:
         print(f"Error occurred while writing data {e}")

def calculate_task_estimation(number_of_task_left, time_for_currect_task):
    total_estimated_time = time_for_currect_task * number_of_task_left
    return f"{round(total_estimated_time/60)} minutes"

def FetchDataFromURL(url):
    try:
        raw_data = requests.get(url)
        print(f"Raw data fetched successfully for url: {url}")
        return raw_data.text
    except Exception as e:
        print(f"There was an error fetching data from provided url {e}")
        quit()


def FetchAllUrlAndUrlIdFromFile(fileName):
    urls_data = []
    data_object = openpyxl.load_workbook(input_file_name)
    sheet = data_object.active
    max_rows = sheet.max_row
    for i in range(2, max_rows + 1):
        urls_data.append(InputFileDataModel(sheet.cell(row = i, column = 1).value, sheet.cell(row = i, column = 2).value))
    return urls_data


def CalculateScoresAndReturnOutputModelData(urlId, url, text, stop_words_from_input, positive_words, negative_words):
   try:
       # Tokenize text into words data
       words = word_tokenize(text)

       lower_case_words = [word.lower() for word in words]
       stop_words = set(stopwords.words('english'))

       # Remove all stopwords
       stop_words_filtered_words = [word for word in words if word.lower() not in stop_words_from_input]

       positive_score = 0
       negative_score = 0
       total_words = len([word.lower() for word in lower_case_words if word.lower() not in stop_words and word not in string.punctuation])
       total_sentences = len(sent_tokenize(text))
       complex_word_count = 0
       personal_pronouns_count = 0
       total_syllables = 0

       #  positive and negative words logic
       for word in stop_words_filtered_words:
           if word in positive_words:
               positive_score += 1
           elif word in negative_words:
               negative_score += 1

       for word in lower_case_words:

           # all personal pronouns count
           if word.lower() in personal_pronouns:
               personal_pronouns_count += 1

           # all syllables count
           word_syllables = sum(word.count(vowel) for vowel in vowels)
           if word.endswith(("es", "ed")):
               word_syllables -= 1

           #complex words count
           if(word_syllables > 2):
                complex_word_count += 1

           total_syllables += max(word_syllables, 1)

       # final score calculation
       polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
       subjectivity_score = (positive_score + negative_score) / (total_words + 0.000001)
       average_sentence_length = total_words / total_sentences
       percentage_complex_words = complex_word_count / total_words
       fog_index = 0.4 * (average_sentence_length + percentage_complex_words)
       average_words_per_sentence = total_words / total_sentences
       average_word_length = sum(len(word) for word in words) / total_words
       syllable_per_word = total_syllables / total_words

       output_file_model = OutputFileModel(urlId, url, positive_score, negative_score, polarity_score, subjectivity_score,
                                           average_sentence_length, percentage_complex_words, fog_index,
                                           average_words_per_sentence, complex_word_count, total_words,
                                           syllable_per_word, personal_pronouns_count, average_word_length)

       return output_file_model
   except Exception as e:
       print(f"There was an error performing data extraction {e}")
       quit()

def CreateAndInsertCalculatedDataToOutputFile(outputFileData):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(output_file_headers)
        for item in outputFileData:
            row_data = [
                item.urlId, item.url, item.positive_score, item.negative_score, item.polarity_score,
                item.subjectivity_score, item.average_sentence_length, item.percentage_complex_words,
                item.fog_index, item.average_words_per_sentence, item.complex_word_count, item.word_count,
                item.syllable_per_word, item.personal_pronouns, item.average_word_length
            ]
            sheet.append(row_data)
        wb.save(output_file_name)
        wb.close()
        print(f"{output_file_name} created successfully!!")
    except PermissionError:
        print("Error occurred while creating output file. Please check close the file and try again!!")
        quit()
    except Exception as e:
        print(f"Error occurred creating the output file {e}")
        quit()

def GetNoDataOutputFile(urlId, url):
    NoData = "No data on website for url"
    return OutputFileModel( urlId, url, NoData, NoData, NoData, NoData, NoData, NoData, NoData, NoData, NoData, NoData, NoData, NoData, NoData)
def ReadFileAndReturnData(fileName, encoding = None):
    encoding_to_use = locale.getpreferredencoding() if encoding is None else encoding
    try:
        with open(fileName, 'r', encoding= encoding_to_use) as file:
            text = file.read()
        print(f"Data read completed successfully for file {fileName}")
        return text
    except Exception as e:
        print(f"There was an error reading the {fileName} file: {e}")

def ReadStopWordsFileAndAppendData():
    appended_stop_words = ""
    try:
        for filename in os.listdir(os.getcwd()):
            if filename.startswith("StopWords_"):
                file_path = os.path.join(os.getcwd(), filename)
                with open(file_path, 'r', encoding= locale.getpreferredencoding()) as file:
                    text = file.read()
                    appended_stop_words += text
        return appended_stop_words
    except Exception as e:
        print(f"There was an error reading all stop words files {e}")
        quit()

def ReadPoisitveWordsFileData():
    return ReadFileAndReturnData(positive_words_file_name)

def ReadNegativeWordsFileData():
    return ReadFileAndReturnData(negative_words_file_name)


if __name__ == "__main__":
    try:
        start_main_time = time.time()
        file_data = FetchAllUrlAndUrlIdFromFile("Input.xlsx")
        print(
            f"\n\nStarting data extraction and file creation process.\nMake sure to include all input files in projects root folder here {os.getcwd()}\n\n")

        total_url_count = len(file_data)
        counter = 0
        CreateFolder(os.path.join(os.getcwd(), extracted_data_folder))
        stop_words_all_file_data = ReadStopWordsFileAndAppendData()
        positive_words_file_data = ReadPoisitveWordsFileData()
        negative_words_file_data = ReadNegativeWordsFileData()
        final_output_file_model_data = []
        for data in file_data:
            counter += 1
            url_id = data.urlId[0]
            raw_data = FetchDataFromURL(data.url)
            extracted_data = ExtractData(raw_data)
            CreateTxtFileAndWriteData(url_id, str(extracted_data))
            extracted_file_data = ReadFileAndReturnData(f"{os.getcwd()}/{extracted_data_folder}/{url_id}.txt", "utf-8")
            if(extracted_file_data == "" or len(extracted_file_data) == 0):
                print(f"No data for {url_id}, {data.url}. Saving No data on website for url in {output_file_name} file as data field.")
                final_output_file_model_data.append(GetNoDataOutputFile(url_id, data.url))
            else:
                final_output_file_model_data.append(
                    CalculateScoresAndReturnOutputModelData(url_id, data.url, extracted_file_data,
                                                            stop_words_all_file_data,
                                                            positive_words_file_data, negative_words_file_data))

            print(f"Process completed {counter}/{total_url_count} \n\n")

        print(f"\n\nCreating Output file data...\n\n")
        # Creating final output file
        CreateAndInsertCalculatedDataToOutputFile(final_output_file_model_data)
        end_main_time = time.time()
        print(f"Data extraction task completed successfully in {round((end_main_time - start_main_time)/60)} minutes.")
    except Exception as e:
        print(f"Program ended with an error {e}")








